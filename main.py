import streamlit as st
import pandas as pd
import re
from io import BytesIO
from pathlib import Path

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =====================================================
# KONFIGURASI HALAMAN
# =====================================================

st.set_page_config(
    page_title="Excel File Merger",
    page_icon="📑",
    layout="wide"
)


st.markdown(
    """
    ### Developed by Prana Ugiana Gio

    **Website:** [pranaugi.com](https://pranaugi.com/)

    **YouTube:** [STATKOMAT](https://www.youtube.com/@STATKOMAT)

    **Online Store:** [lynk.id/statkomat](https://lynk.id/statkomat)

    **Training Data for This Application:**  
    [Download training data from Google Drive](https://drive.google.com/drive/folders/1aMNWaiBNt-qrsCYOKROIJ9XNon4A33mq?usp=sharing)


    ---
    """
)





st.title("📑 Excel / CSV File Merger")
st.caption(
    "Upload beberapa file Excel atau CSV, lalu aplikasi akan menggabungkan data "
    "secara vertikal / ke bawah dan mengekspor hasilnya ke Excel."
)


# =====================================================
# FUNGSI BANTUAN
# =====================================================

def extract_code_from_filename(filename):
    """
    Contoh:
    AADI 2022-2025.xlsx -> AADI
    ABMM 2022-2025.xlsx -> ABMM
    """

    name = Path(filename).stem
    match = re.match(r"([A-Za-z0-9]+)", name)

    if match:
        return match.group(1).upper()

    return name


def get_excel_sheet_names(uploaded_file):
    file_bytes = uploaded_file.getvalue()
    xls = pd.ExcelFile(BytesIO(file_bytes))
    return xls.sheet_names


def read_uploaded_file(uploaded_file, sheet_option="Sheet pertama"):
    """
    Membaca file Excel / CSV dari Streamlit uploader.
    """

    filename = uploaded_file.name
    suffix = Path(filename).suffix.lower()
    file_bytes = uploaded_file.getvalue()

    if suffix == ".csv":
        try:
            df = pd.read_csv(BytesIO(file_bytes))
        except UnicodeDecodeError:
            df = pd.read_csv(BytesIO(file_bytes), encoding="latin-1")

        return df

    elif suffix in [".xlsx", ".xls"]:
        xls = pd.ExcelFile(BytesIO(file_bytes))
        sheet_names = xls.sheet_names

        if sheet_option == "Sheet pertama":
            sheet_name = sheet_names[0]
        else:
            sheet_name = sheet_option

            if sheet_name not in sheet_names:
                sheet_name = sheet_names[0]

        df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)

        return df

    else:
        raise ValueError(f"Format file tidak didukung: {filename}")


def clean_dataframe(df):
    """
    Membersihkan dataframe:
    - Menghapus baris yang seluruhnya kosong
    - Menghapus kolom Unnamed kosong
    - Merapikan nama kolom
    """

    df = df.copy()

    # Rapikan nama kolom
    df.columns = df.columns.astype(str).str.strip()

    # Hapus kolom Unnamed yang seluruhnya kosong
    unnamed_cols = [
        col for col in df.columns
        if col.lower().startswith("unnamed")
    ]

    for col in unnamed_cols:
        if df[col].isna().all():
            df = df.drop(columns=[col])

    # Hapus baris yang seluruhnya kosong
    df = df.dropna(how="all")

    return df


def combine_files(
    uploaded_files,
    sheet_option,
    column_mode,
    add_source_file,
    add_source_code
):
    """
    Menggabungkan beberapa file secara vertikal.
    """

    dataframes = []
    log_rows = []

    for file in uploaded_files:
        try:
            df = read_uploaded_file(file, sheet_option=sheet_option)
            df = clean_dataframe(df)

            if add_source_code:
                df.insert(0, "Source Code", extract_code_from_filename(file.name))

            if add_source_file:
                df.insert(0, "Source File", file.name)

            dataframes.append(df)

            log_rows.append({
                "File": file.name,
                "Status": "OK",
                "Rows": len(df),
                "Columns": len(df.columns),
                "Message": "Berhasil dibaca"
            })

        except Exception as e:
            log_rows.append({
                "File": file.name,
                "Status": "ERROR",
                "Rows": 0,
                "Columns": 0,
                "Message": str(e)
            })

    if not dataframes:
        return pd.DataFrame(), pd.DataFrame(log_rows)

    if column_mode == "Gabungkan semua kolom":
        combined_df = pd.concat(
            dataframes,
            ignore_index=True,
            sort=False
        )

    else:
        # Ambil hanya kolom yang sama di semua file
        common_columns = set(dataframes[0].columns)

        for df in dataframes[1:]:
            common_columns = common_columns.intersection(set(df.columns))

        common_columns = list(dataframes[0].columns.intersection(common_columns))

        dataframes_common = [
            df[common_columns].copy()
            for df in dataframes
        ]

        combined_df = pd.concat(
            dataframes_common,
            ignore_index=True
        )

    return combined_df, pd.DataFrame(log_rows)


def dataframe_to_excel_bytes(df, log_df=None):
    """
    Menyimpan hasil gabungan ke Excel.
    Sheet utama: Combined Data
    Sheet log: Log
    """

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Combined Data", index=False)

        if log_df is not None and not log_df.empty:
            log_df.to_excel(writer, sheet_name="Log", index=False)

        workbook = writer.book
        worksheet = writer.sheets["Combined Data"]

        # Style
        header_fill = PatternFill(
            start_color="D9EAF7",
            end_color="D9EAF7",
            fill_type="solid"
        )

        thin_border = Border(
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
            top=Side(style="thin", color="DDDDDD"),
            bottom=Side(style="thin", color="DDDDDD")
        )

        # Freeze header
        worksheet.freeze_panes = "A2"

        # Format header
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.border = thin_border

        # Format isi
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        # Lebar kolom otomatis sederhana
        for col_idx in range(1, worksheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0

            for cell in worksheet[col_letter]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max(max_length + 2, 12), 35)
            worksheet.column_dimensions[col_letter].width = adjusted_width

        # Format sheet Log
        if log_df is not None and not log_df.empty:
            log_sheet = writer.sheets["Log"]
            log_sheet.freeze_panes = "A2"

            for cell in log_sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                cell.border = thin_border

            for row in log_sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = thin_border

            for col_idx in range(1, log_sheet.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0

                for cell in log_sheet[col_letter]:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))

                adjusted_width = min(max(max_length + 2, 12), 50)
                log_sheet.column_dimensions[col_letter].width = adjusted_width

    output.seek(0)
    return output.getvalue()


# =====================================================
# SIDEBAR UPLOAD
# =====================================================

st.sidebar.header("1. Upload File")

uploaded_files = st.sidebar.file_uploader(
    "Upload beberapa file Excel atau CSV",
    type=["xlsx", "xls", "csv"],
    accept_multiple_files=True
)

st.sidebar.header("2. Pengaturan Gabungan")

sheet_option = st.sidebar.text_input(
    "Nama sheet Excel yang dibaca",
    value="Sheet pertama",
    help=(
        "Isi 'Sheet pertama' untuk membaca sheet pertama. "
        "Atau isi nama sheet tertentu jika semua file memiliki nama sheet yang sama."
    )
)

column_mode = st.sidebar.radio(
    "Mode penggabungan kolom",
    [
        "Gabungkan semua kolom",
        "Hanya kolom yang sama di semua file"
    ],
    index=0
)

add_source_file = st.sidebar.checkbox(
    "Tambahkan kolom Source File",
    value=True
)

add_source_code = st.sidebar.checkbox(
    "Tambahkan kolom Source Code dari nama file",
    value=True
)

process_button = st.sidebar.button(
    "🚀 Gabungkan File",
    type="primary"
)


# =====================================================
# MAIN APP
# =====================================================

if not uploaded_files:
    st.info("Silakan upload beberapa file Excel atau CSV terlebih dahulu.")
    st.stop()

st.subheader("File yang Diupload")

uploaded_info = pd.DataFrame({
    "No": range(1, len(uploaded_files) + 1),
    "File Name": [file.name for file in uploaded_files],
    "File Size KB": [round(len(file.getvalue()) / 1024, 2) for file in uploaded_files]
})

st.dataframe(uploaded_info, use_container_width=True)

if not process_button:
    st.info("Klik tombol **Gabungkan File** untuk mulai memproses data.")
    st.stop()


# =====================================================
# PROSES GABUNGAN
# =====================================================

with st.spinner("Sedang membaca dan menggabungkan file..."):
    combined_df, log_df = combine_files(
        uploaded_files=uploaded_files,
        sheet_option=sheet_option,
        column_mode=column_mode,
        add_source_file=add_source_file,
        add_source_code=add_source_code
    )

if combined_df.empty:
    st.error("Tidak ada data yang berhasil digabungkan.")
    st.dataframe(log_df, use_container_width=True)
    st.stop()


# =====================================================
# PREVIEW HASIL
# =====================================================

st.success("File berhasil digabungkan.")

col1, col2, col3 = st.columns(3)

with col1:
    st.metric("Jumlah File Diupload", len(uploaded_files))

with col2:
    st.metric("Jumlah Baris Gabungan", len(combined_df))

with col3:
    st.metric("Jumlah Kolom Gabungan", len(combined_df.columns))


st.subheader("Preview Data Gabungan")
st.dataframe(combined_df.head(100), use_container_width=True)

st.subheader("Log Proses")
st.dataframe(log_df, use_container_width=True)


# =====================================================
# DOWNLOAD EXCEL
# =====================================================

excel_bytes = dataframe_to_excel_bytes(
    df=combined_df,
    log_df=log_df
)

st.download_button(
    label="⬇️ Download Excel Hasil Gabungan",
    data=excel_bytes,
    file_name="combined_data.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)